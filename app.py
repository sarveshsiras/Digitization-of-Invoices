from flask import Flask,make_response,request,jsonify,render_template,url_for,redirect,flash
from forms import LoginForm,CreateAccountForm
from flask_mongoengine import MongoEngine
from config import DB_connection
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager,login_user,login_required,logout_user,current_user
#import User
import openpyxl
from openpyxl.styles import Alignment
from flask import send_from_directory
import os
from werkzeug.utils import secure_filename
from flask_login import UserMixin
from s3_upload import upload_to_s3
import pandas as pd
import interface

app = Flask(__name__)
app.config['SECRET_KEY'] = 'the random string'  
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
#db.init_app(app)
conn = DB_connection()
DB_URI = "mongodb+srv://{}:{}@cluster0.1aksc.mongodb.net/{}?retryWrites=true&w=majority".format(conn.user_name,conn.password,conn.database_name)

app.config["MONGODB_HOST"] = DB_URI
db = MongoEngine(app)

UPLOAD_FOLDER = 'static/uploads/'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

class user_data(db.Document, UserMixin):

    username = db.StringField()
    email = db.StringField()
    password = db.StringField()
    uploaded_image_urls = db.ListField()
    processed_excel_urls = db.ListField()

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_invoice(URL, detected_table,returned_coordinates):
    processed_data = interface.table_processing_interface(detected_table)
    non_tabular_data = interface.non_tabular_data_interface(URL, returned_coordinates)
    return processed_data,non_tabular_data

def write_in_excel(processed_data, non_tabular_data, selected):
    if(selected == None):
        print("HAHAHA--------------------------------")
        wb = openpyxl.Workbook()
        sheet = wb.active
        le_ = len(processed_data)
        p = 0
        for i in processed_data:
            k = 0
            for j in i:
                temp = j.split("\x0c")
                print(temp)
                c1 = sheet.cell(row=p+1,column=k+1)
                c1.value = str(temp[0].rstrip("\n"))
                k+=1
            p+=1
        wb.save("demo1.xlsx")
        bianry_data = open("demo1.xlsx","rb")
        filename = extract_file_names(current_user.uploaded_image_urls[-1]).split(".")[0] + ".xlsx"
        excel_url = upload_to_s3(file_data=bianry_data, username = current_user.username, filename = filename)
        current_user.processed_excel_urls.append(excel_url)
        current_user.save()
        print(excel_url)
        return excel_url
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        le_ = len(processed_data)
        p = 0
        for i in processed_data:
            k = 0
            for j in i:
                temp = j.split("\x0c")
                print(temp)
                c1 = sheet.cell(row=p+1,column=k+1)
                c1.value = str(temp[0].rstrip("\n"))
                k+=1
            p+=1
        #wb.save("demo1.xlsx")
        #sheet.autoFitRow(p+3)
        #sheet.row_dimensions[p+3].width = 400
        sheet.cell(row = p + 2, column = 1).value= "Non-Tabular Data"

        sheet.cell(row = p + 3, column = 1).value= non_tabular_data.strip('\n').split("\x0c")[0]

        sheet.cell(row = p + 3, column = 1).alignment = Alignment(wrap_text=True)

        sheet.column_dimensions['A'].width = 30

        sheet.row_dimensions[p+3].height = 4000

        print("Success")
        wb.save("demo1.xlsx")
        bianry_data = open("demo1.xlsx","rb")
        filename = extract_file_names(current_user.uploaded_image_urls[-1]).split(".")[0] + ".xlsx"
        excel_url = upload_to_s3(file_data=bianry_data, username = current_user.username, filename = filename)
        print(excel_url)
        current_user.processed_excel_urls.append(excel_url)
        current_user.save()
        return excel_url

def get_row_values(dictionary):
    keys = list(dictionary.keys())
    row_values = []
    for i in range(1,len(keys) + 1):
        key = str(i)
        values = []
        flag = 1
        for j in keys:
            key_num = j.split("-")[0]
            if(key_num == key):
                flag = 0
                values.append(dictionary[j])
            elif(flag == 0):
                break
        if(len(values) != 0):
            row_values.append(values)
        else:
            break
    return row_values



@app.route("/invoice_data", methods = ['GET','POST'])
@login_required
def get_invoice_data():
    URL = current_user.uploaded_image_urls[-1]
    print(URL)
    #processed_data, non_tabular_data = process_invoice(URL)
    if(request.method == 'GET'):
        return render_template("invoice_data.html",invoice_url = URL, data = processed_data, non_tabular_data = non_tabular_data)
    else:
        row_values = []
        value = request.form.get('include_non_tabular_data')
        print(value)
        non_tabular_data = request.form['Non_tabular_Data']
        print(non_tabular_data)
        print("invodata-------------")
        f = request.form
        print("f", f)
        values = get_row_values(f)
        print(values)
        downloadable_link = write_in_excel(values,non_tabular_data,value)
        return render_template("download_excel.html",downloadlink = downloadable_link)


@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    
    return redirect(url_for('static',filename='uploads/' + filename), code=301)

@app.route('/upload_invoice',methods=['GET','POST'])
@login_required
def upload_file():
    if request.method == 'POST':
        print("INSIDE")
        # check if the post request has the file part
        if 'file1' not in request.files:
            flash('No file part')
            return redirect(url_for('upload'))
        file = request.files['file1']
        # if user does not select file, browser also
        # submit a empty part without filename
        print(file.filename)
        if file.filename == '':
            flash('No selected file')
            return redirect(url_for('upload_invoice'))
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            #file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            URL = upload_to_s3(file_data=file, username = current_user.username, filename = file.filename)
            # URL = current_user.uploaded_image_urls[-1]
            current_user.uploaded_image_urls.append(URL)
            current_user.save()
            return render_template('upload.html', filename=URL)
        else:
            flash('Select proper format for the image')
            return redirect(url_for('upload'))
    else:
        return render_template('upload.html')


def extract_file_names(name):
    rev = name.split("___")[1]
    return rev

@app.route('/crop', methods=['GET', 'POST'])
def get_table():
    URL = current_user.uploaded_image_urls[-1]
    returned_coordinates = interface.get_table_coordinates_interface(URL)
    return render_template('crop.html', filename=URL, returned_coordinates = returned_coordinates)

@app.route('/dashboard')
@login_required
def dashboard():
    uploaded_url_list = current_user.uploaded_image_urls
    print(uploaded_url_list)
    number_of_invoices = len(uploaded_url_list)
    file_name_list = []
    for url in uploaded_url_list:
        file_name_list.append(extract_file_names(url))
    print(number_of_invoices)
    print(file_name_list)
    processed_excel_urls = current_user.processed_excel_urls
    uploaded_urls = zip(uploaded_url_list,file_name_list,processed_excel_urls)
    return render_template('dashboard.html', segment='index', no_of_invoices_processed = number_of_invoices, url_list = uploaded_urls)


@app.route('/profile' ,methods = ['GET','POST'])
@login_required
def profile():
    number_of_invoices = len(current_user.uploaded_image_urls)
    if(request.method == 'GET'):
        return render_template('profile.html',Number_of_invoices = number_of_invoices, pass_msg = "")
    else:
        password_1 = request.form['Password_1']
        password_2 = request.form['Password_2']
        if(password_1 != password_2):
            return render_template('profile.html',Number_of_invoices = number_of_invoices, pass_msg_1 = "Passwords don't match. Please Reenter",pass_msg_2 = "")

        else:
            hashpass = generate_password_hash(password_1, method='sha256')
            current_user.password = hashpass
            current_user.save()
            return render_template('profile.html',Number_of_invoices = number_of_invoices, pass_msg_1 = "",pass_msg_2 = "Password Saved Successfully")

@app.route('/')
def default():
    return redirect(url_for('login'))

@app.route("/login",methods = ['GET','POST'])
def login():
    login_form = LoginForm(request.form)
    if 'login' in request.form:
        
        # read form data
        username = request.form['username']
        password = request.form['password']

        #Locate user
        logged_in_user = user_data.objects(username=username).first()

        # Check the password
        if logged_in_user and check_password_hash(logged_in_user.password, password):
            print("YES")
            login_user(logged_in_user)
            return redirect(url_for('default'))

        # Something (user or pass) is not ok
        return render_template( 'accounts/login.html', msg='Wrong Username or password', form=login_form)

    if not current_user.is_authenticated:
        return render_template( 'accounts/login.html',
                                form=login_form)
    
    return redirect(url_for('dashboard'))

@app.route('/register',methods=['GET','POST'])
def register():
    form = LoginForm(request.form)
    create_account_form = CreateAccountForm(request.form)
    if 'register' in request.form:

        username  = request.form['username']
        email     = request.form['email'   ]
        password  = request.form['password']


        if form.validate():
            # Check usename exists
            existing_username = user_data.objects(username=username).first()
            #print(existing_username.username)
            if existing_username:
                return render_template( 'accounts/register.html', 
                                        msg='Username already registered',
                                        success=False,
                                        form=create_account_form)
            
            existing_mail = user_data.objects(email=email).first()
            if existing_mail:
                return render_template( 'accounts/register.html', 
                                        msg='Email already registered',
                                        success=False,
                                        form=create_account_form)


            else:
                hashpass = generate_password_hash(password, method='sha256')
                new_user = user_data(username = username,email = email,password = hashpass)
                new_user.save()
                login_user(new_user)
                return render_template( 'accounts/register.html', 
                                msg='User created please <a href="/login">login</a>', 
                                success=True,
                                form=create_account_form) 

    else:
        return render_template( 'accounts/register.html', form=create_account_form)


@app.route('/display', methods=['POST'])
@login_required
def show():
    if request.method == 'POST':
        x1 = request.form['X1']
        y1 = request.form['Y1']
        x2 = request.form['X2']
        y2 = request.form['Y2']
        print(x1, y1, x2, y2)
        returned_coordinates = [x1, y1, x2, y2]
        URL = current_user.uploaded_image_urls[-1]
        detected_table = interface.crop_image_interface(URL, returned_coordinates)
        text_data, non_tabular_data = process_invoice(URL, detected_table, returned_coordinates)
        return render_template("invoice_data.html",invoice_url = URL, data = text_data, non_tabular_data = non_tabular_data)

@app.route('/render', methods=['POST'])
@login_required
def get_names():
    if request.method == 'POST':
        boxes = request.get_json()
        print(boxes['boxes'][0].values())

    return redirect(url_for("index"))

@app.route('/logout')
@login_required
def logout():
    login_form = LoginForm(request.form)
    logout_user()
    print("WORKS")
    return redirect(url_for('login'))

@login_manager.user_loader
def load_user(user_id):
    return user_data.objects(pk=user_id).first()

if __name__ == "__main__":
    app.run(debug=True)